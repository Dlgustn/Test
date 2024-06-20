import requests
import time # 딜레이 라이브러리
from selenium import webdriver
from selenium.webdriver import ChromeOptions
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.support.select import Select
from selenium.webdriver.common.by import By # 셀레니움 4.3 이상부터 동작 방식이 변경됨에 따라 By 사용법 숙지 필수
import webdriver_manager
from selenium.webdriver.support import expected_conditions as EC
import openpyxl
from os import path
import clipboard # 클립보드 라이브러리

#엑셀에 있는 회사명을 읽어와 네이버 지도에 검색 후 각 회사의 도로명 주소 가져오기 

#----------------- 엑셀 파일 열기 + 엑셀에 있는 값(회사명) 가져오기 ---------------------

#어떤 컴퓨터에서도 쓸 수 있게 환경 변수 가져오기
def find_roadname():
    username = path.expanduser('~')

    #파일 경로 할당
    file_name = username + "\Desktop\PythonTest.xlsx"

    #파일 열기(이미 존재하는 파일)
    wb = openpyxl.load_workbook(file_name)

    #엑셀 파일 경로 출력(확인용)
    print(file_name)

    #엑셀 파일에서 시트 이름들 가져오기
    sheetlist = wb.sheetnames

    #시트2 활성화
    ws = wb['Sheet2']

    #활성화된 시트 가져오기
    sheet2 = wb.active

    #시트 이름 할당
    sheet_title = sheet2.title

    #활성화된 시트 이름 출력(확인용)
    print(sheet_title)

    #시트에 있는 A컬럼 값들 리스트로 가져오기
    companynamelist=[]
    for row in sheet2.rows:
        companynamelist.append(
            row[0].value
        )

    #엑셀에서 가져온 리스트 출력(확인용)    
    print(companynamelist)

    #엑셀에서 가져온 리스트 길이 할당
    excellistlen = len(companynamelist)

    #확인용 출력
    print(excellistlen)

    """
    c=[]
    c.append(sheet2.cell(row=1,column=1).value)
    c.append(sheet2.cell(row=2,column=1).value)
    c.append(sheet2.cell(row=3,column=1).value)

    """

    #----------------- 웹브라우저 열기에 필요한 준비 ---------------------

    #셀레니움 라이브러리에서 크롬 웹드라이버 설정 가져오기
    options = webdriver.ChromeOptions()
    #options = options()

    #브라우저 창 최대화
    options.add_argument("--start-maximized")

    #크롬 웹드라이버 옵션 적용한 객체 생성(? gweb 비슷한거? 크롬창 열림)
    driver = webdriver.Chrome(options)

    #a=["인지소프트", "안랩", "토스"]

    #----------------- 네이버 지도 접속 및 도로명 주소 가져오기 ---------------------

    #도로명 주소 담을 리스트 생성
    roadnamelist = []

    #각 회사별 네이버 지도 검색결과 들어가서 도로명 주소 가져오기
    for i in range(0,excellistlen, 1):
        
        #회사 이름 할당
        companyname = companynamelist[i]
        #print(companyname)

        #공통 url 할당
        url = "https://map.naver.com/p/search/"
        
        #회사 url 할당
        companyurl = url + str(companyname)
        
        #implicitly_wait(5)는 암시적 대기 즉 최대 5초 기다리되 5초 안에 작업이 완료되면 더 기다리지 않음
        
        #url로 이동
        driver.get(companyurl)
        driver.implicitly_wait(5)

        #iframe 벗기기
        driver.switch_to.frame("searchIframe")
        
        #Xpath 활용해서 도로명 주소 가져오기
        #roadname = driver.find_element(By.XPATH, '//*[@id="_pcmap_list_scroll_container"]/ul/li[1]/div[1]/div/div/div/span[2]/a/span[1]')
        
        try:
            driver.find_element(By.XPATH, '//*[@id="app-root"]/div/div[2]/div/div')
            roadnamelist.append('검색결과가 없습니다.')
        except:
            #Xpath 활용해서 도로명 주소 나오게 하는 버튼 클릭
            driver.find_element(By.XPATH, '//*[@id="_pcmap_list_scroll_container"]/ul/li[1]/div[1]/div/div/div/span[2]/a/span[1]').click()

            #Xpath 활용해서 복사 버튼 클릭
            driver.find_element(By.XPATH, '//*[@id="_pcmap_list_scroll_container"]/ul/li[1]/div[1]/div/div/div/div/div/div[1]/span[2]/a').click()
            
            #클립보드에 저장된 값 할당
            clip_df = clipboard.paste()
            #print(clip_df)

            #도로명 주소 리스트에 담기
            roadnamelist.append(clip_df)


    # 결과 확인
    print(*roadnamelist, sep='\n')

    wb.close

    #파일 열기(이미 존재하는 파일)
    wb = openpyxl.load_workbook(file_name)

    #엑셀 파일에서 시트 이름들 가져오기
    sheetlist = wb.sheetnames

    #시트2 활성화
    ws = wb['Sheet2']

    #활성화된 시트 가져오기
    sheet2 = wb.active

    #시트에 있는 B컬럼에 도로명 주소들 입력
    for i in range(0,excellistlen, 1):
        rownum = i + 1
        sheet2.cell(row=rownum,column=2).value = roadnamelist[i]
        
    wb.save(file_name)
    print("성공")
    return("성공")

find_roadname()