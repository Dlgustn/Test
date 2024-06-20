"""
import openpyxl
import pandas as pd
from os import path

username = path.expanduser('~')

#파일명
file_name = username + "\Desktop\PythonTest.xlsx"

#데이터 프레임 형식으로 엑셀 읽기
df = pd.read_excel(file_name, engine='openpyxl')
#출력
print(df)

wb = openpyxl.load_workbook(file_name)
sheet1 = wb.active
sheet_title = sheet1.title
print(sheet_title)

sheet1.cell(row=4,column=1).value="iauto"
sheet1.cell(row=4,column=2).value="이현수"
wb.save(file_name)

row = 5
newrow = row * 2
for i in range(newrow, 1, -2):
    spacenum = i // 2
    space = spacenum  * ' '
    starcount = newrow - i + 1
    star = starcount * '*'
    print(space + star + space)
"""

# a = 123
# c = '123'
# print(type (a), type (c))

# d = c + str(a)
# print(d)

# for i in range(5):
#     print(i)

# for i in range(1,5,1): 
#      print(i)

# list = [1,2,"a",4,5]

# for i in list:
#     if i == 1:
#         print(i)

# a = 1
# while True:
#     a = 2
#     print(a)
#     break
# print(a)

# a = "    abc    "
# a = a.strip()
# print(a)

#피라미드 찍기

# for i in range(1,10,2):
#     star = '*' * i
#     for j in range(8,2,-2):
#         k = j / 2
#         space = k * ' '
#         print(space + star + space)



import openpyxl
import pandas as pd
from os import path

username = path.expanduser('~')

#파일명
file_name = username + "\Desktop\PythonTest.xlsx"

"""
#df = pd.read_excel(file_name, engine='openpyxl', sheet_name='Sheet2')
#print(df)

#파일 열기
wb = openpyxl.load_workbook(file_name)

#파일 이름 출력(확인용)
print(file_name)

#열린파일에서 시트 네임들 가져오기
wb.sheetnames

#시트2 활성화
ws = wb['Sheet2']

#활성화된 시트 가져오기
sheet2 = wb.active

#활성화된 시트 이름 출력(확인용)
sheet_title = sheet2.title
print(sheet_title)

#셀 위치로 피라미드 로우 값 가져오기
row = sheet2.cell(row=1,column=1).value

#로우 값 출력(확인용)
print(row)
#wb.save(file_name)

#row = 5
newrow = row * 2
for i in range(newrow, 1, -2):
    spacenum = i // 2
    space = spacenum  * ' '
    starcount = newrow - i + 1
    star = starcount * '*'
    print(space + star + space)
"""

#파일 열기
wb = openpyxl.load_workbook(file_name)

#파일 이름 출력(확인용)
print(file_name)

#열린파일에서 시트 네임들 가져오기
a = wb.sheetnames
print(a)

#시트2 활성화
ws = wb['Sheet2']

#활성화된 시트 가져오기
sheet2 = wb.active

#활성화된 시트 이름 출력(확인용)
sheet_title = sheet2.title
print(sheet_title)

#셀 위치로 피라미드 로우 값 가져오기
sheet2.cell(row=1,column=1).value = "저장을 했다."
wb.save(file_name)

row = sheet2.cell(row=1,column=1).value

print(row)
openpyxl.__version__


