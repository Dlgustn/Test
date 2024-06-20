import openpyxl
from os import path

def Temp():
  username = path.expanduser('~')

  #파일명
  file_name = username + "\\Desktop\\PythonTest.xlsx"
  print(file_name)

  #파일 열기
  wb = openpyxl.load_workbook(file_name)

  #파일 이름 출력(확인용)
  #print(file_name)

  #열린파일에서 시트 네임들 가져오기
  wb.sheetnames

  #시트2 활성화
  ws = wb['Sheet2']

  #활성화된 시트 가져오기
  sheet2 = wb.active

  #활성화된 시트 이름 출력(확인용)
  sheet_title = sheet2.title
  #print(sheet_title)

  a = sheet2.cell(row=1,column=1).value
  print(str(a)+"장형렬")
  return str(a) + "이현수"

Temp()


