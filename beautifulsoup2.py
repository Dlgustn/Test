import requests
# html을 받아오기 위한 요청을 보낼 때 쓰는 라이브러리

from bs4 import BeautifulSoup
# 백그라운드로 html을 받아서 파싱할 때 쓰는 라이브러리

import openpyxl
# 엑셀 여는 라이브러리

from os import path
# 환경 변수(파일경로) 위한 라이브러리

import time
# 시간 딜레이(아이오토로는 타임슬립) 라이브러리

from urllib.parse import quote
# url 중 회사명 인코딩 하는 라이브러리

import pandas as pd
# 엑셀 데이터를 읽어오기 위한 라이브러리


def temp():
    # 엑셀 파일 경로 할당
    userName = path.expanduser('~')
    #file_name = username + "\Desktop\PythonTest.xlsx"
    
    fileName = path.join(userName, "Desktop", "PythonTest.xlsx")
    # 파일 경로 조합 및 할당
    
    print(fileName)

    # 엑셀 파일 열기
    wb = openpyxl.load_workbook(fileName)

    # 시트 리스트 받아오기
    sheetList = wb.sheetnames
    
    # 시트 선택
    ws = wb['Sheet2']
    
    # 선택한 시트 활성화
    sheet2 = wb.active

    # 시트 타이틀 할당(확인용)
    sheetTitle = sheet2.title
    print(sheetTitle)

    # 회사 이름들 할당할 리스트 생성
    companyNameList = []

    # 시트에서 회사 이름들 읽어오기 > 각 로우마다 있으면 회사이름 리스트에 추가
    for aRow in sheet2.iter_rows(min_row=1, max_col=1, values_only=True):
        companyName = aRow[0]
        if companyName is not None:
            companyNameList.append(companyName)

    # 회사 이름 리스트 길이 할당 > 회사 수 만큼 검색해야 하므로 추후 for문에 활용
    excellistlen = len(companyNameList)
    print(companyNameList)

    """
    for Brow in sheet2.iter_rows(min_row=2, max_col=1, values_only=True):
        companyName = Brow[0]
        if companyName is not None:
            companyNameList.append(companyName)
    """


    # url에서 html 받아오는 함수 재시도 회수 3 실패 시 2배씩 대기 시간 증가(최초 1초)
    def getPageContent(url, retries=3, backoffFactor=1.0):
        
        # 해더 할당 == 로봇이 아닌 것 처럼 보이게 함 > 사람이 브라우저 여는 것 처럼 
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
        }
        
        for i in range(retries):
            
            # html 요청
            response = requests.get(url, headers=headers)
            
            # 응답 받기 성공이면 인코딩 후 리턴
            if response.status_code == 200:
                response.encoding = 'utf-8'
                return response.text
            
            # 응답 못받으면 대기 후 재시도
            elif response.status_code == 429:
                print(f"Too many requests. Waiting {backoffFactor} seconds before retrying...")
                time.sleep(backoffFactor)
                backoffFactor *= 2

            # 다른 이유로 실패 시 로그 찍기
            else:
                print(f"Failed to retrieve the webpage. Status code: {response.status_code}")
                break
        
        return None

    #html에서 첫번째 검색결과의 도로명 주소 받아오기

    # 도로명 주소 저장할 리스트 할당
    roadNameList = []

    # 공통 url 할당
    url = "https://pcmap.place.naver.com/mart/list?query="

    for companyName in companyNameList:
        print(companyName)
        
        # 회사명 인코딩해서 url 조합
        encodedCompanyName = quote(companyName)
        companyurl = url + encodedCompanyName

        # url 매개변수로 get_page_content 함수 수행 > html 받아와서 html_content에 할당
        htmlContent = getPageContent(companyurl)

        # html 있으면 뷰티플수프 이용해서 파싱
        if htmlContent:
            soup = BeautifulSoup(htmlContent, 'lxml')

            # 파싱한 html에서 태그, 클래스로 요소들(회사 검색 결과들) 찾기
            spanElements = soup.find_all('span', class_='Pb4bU')

            # 검색 결과들이 있으면 그중 가장 첫 검색 결과의 도로명 주소 받아오기
            if spanElements:
                firstSpanText = spanElements[0].get_text()  # 첫 번째 span 요소의 텍스트만 가져옴
                roadNameList.append(firstSpanText) 

            # 검색결과 없으면 로그 찍고 "검색 결과가 없습니다." 리스트에 추가
            else:
                print("Element not found")
                roadNameList.append("검색 결과가 없습니다.")
        
        # html이 없으면(못 받아오면) == 응답(response) 못 받으면 로그
        else:
            print("Failed to retrieve the webpage after retries.")

    print(roadNameList)

    # 엑셀에 도로명 주소 쓰고 저장
    wb = openpyxl.load_workbook(fileName)
    ws = wb['Sheet2']

    # roadname_list를 순서대로 B1부터 쓰기
    for i, roadName in enumerate(roadNameList, start=1):
        cell = ws[f'B{i}']
        cell.value = roadName

    # 엑셀 파일 저장
    wb.save(fileName)

    # 엑셀 파일 닫기
    wb.close()


temp()






#2 웹 없이 가져오기

# 엘리먼트 찾기
#elements = soup.find_all('div', {'class': 'Pb4bU'})
#element = soup.select_one('#_pcmap_list_scroll_container > ul > li:nth-child(1) > div > div:nth-child(2) > div > span:nth-child(2) > a > span:nth-child(1)')



# 텍스트 출력
#partial_html = soup.select_one("#_pcmap_list_scroll_container")

#partial_html = soup.find(By.XPATH, '//*[@id="_pcmap_list_scroll_container')

#if partial_html:
#    print(partial_html.prettify())
#else:
#    print("해당 요소를 찾을 수 없습니다.")
#element = soup.select_one("#_pcmap_list_scroll_container > ul > li:nth-child(1) > div:nth-child(1) > div > div > div > span:nth-child(2) > a > span:nth-child(1)")
#element = soup.select_one("div.Uao0X a span.Pb4bU")
#element = soup.select_one("div.Uao0X a span.Pb4bU")
#element = soup.find(By.XPATH, '//*[@id="_pcmap_list_scroll_container"]/ul/li[1]/div[1]/div/div/div/span[2]/a/span[1]')

"""
if element:
    print(element.text)
else:
    print("요소를 찾을 수 없습니다.")
"""
#element = soup.select_one('#_pcmap_list_scroll_container > ul > li:nth-child(1) > div.RZdkC > div > div > div > span:nth-child(2) > a > span')
#print(element)